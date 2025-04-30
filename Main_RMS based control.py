import numpy as np
import pygame
import pygame.camera
import cv2
import serial
import math
import tensorflow as tf
from pygame import Color
from PIL import Image
import time
from yolo import YOLO
from unet import Unet
from collections import deque
import openpyxl
import math
from RMS.CCClient import CCClient

# =================================
# GPU Configuration for TensorFlow
# =================================
gpus = tf.config.experimental.list_physical_devices(device_type='GPU')
for gpu in gpus:
    tf.config.experimental.set_memory_growth(gpu, True)

# ===========================
# Excel Workbook Initialization
# ===========================
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = 'pos.xlsx'

wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'pos_next.xlsx'

# =========================== #
# Robot Communication Setup   #
# =========================== #
cps = CCClient()
cps.connectTCPSocket('192.168.0.10')
cps.SetOverride(0.01)
fname = "servoPdata.txt"

# ===========================
# Global Variables
# ===========================
last_pos = (0, 0)
begin_pos = (0, 0)
end_pos = (0, 10)
pos = []
pos_next = []
flag = True

# ===========================
# Utility Functions
# ===========================
def render(route, frame):
    for i in range(len(route) - 1):
        x1, y1 = route[i]
        x2, y2 = route[i + 1]
        cv2.line(frame, (y1, x1), (y2, x2), (0, 0, 255), 1)
    cv2.imshow("video1", frame)
    output_video.write(frame)
    cv2.waitKey(1)

def BFS_choice(cur, Matrix, front, visit, q):
    length = Matrix.shape[0]
    width = Matrix.shape[1]
    directions = [(1, 0), (-1, 0), (0, 1), (0, -1), (1, 1), (-1, -1), (1, -1), (-1, 1)]
    for dx, dy in directions:
        new_x = cur % length + dx
        new_y = cur // length + dy
        if 0 <= new_x < length and 0 <= new_y < width:
            new_index = new_x + new_y * length
            if visit[new_index] and not Matrix[new_x][new_y]:
                front[new_index] = cur
                visit[new_index] = 0
                q.append(new_index)

def BFS(Matrix, begin_pos, target, length, front, visit):
    target_flag = False
    target_index = target[0] + target[1] * Matrix.shape[0]
    begin_index = begin_pos[0] + begin_pos[1] * Matrix.shape[0]
    q = deque()
    q.append(begin_index)
    visit[begin_index] = 0
    while q:
        cur = q.popleft()
        if target_index == cur:
            target_flag = True
            break
        BFS_choice(cur, Matrix, front, visit, q)
    return target_flag

def bezier_curve(points, n_points=100):
    def bernstein_poly(i, n, t):
        return math.comb(n, i) * (t ** i) * ((1 - t) ** (n - i))

    def bezier_interp(points, t):
        n = len(points) - 1
        x, y = 0, 0
        for i in range(n + 1):
            bern = bernstein_poly(i, n, t)
            x += points[i][0] * bern
            y += points[i][1] * bern
        return int(x), int(y)
    curve = []
    for step in range(n_points):
        t = step / (n_points - 1)
        curve.append(bezier_interp(points, t))
    return curve

def compute_P1():
    return (int((begin_pos[0] + end_pos[0]) / 2 + 45), int((begin_pos[1] + end_pos[1]) / 2 + 45))

def event_handler(event, x, y, flags, param):
    global begin_pos, end_pos, flag, max_distance
    if event == cv2.EVENT_RBUTTONDOWN:
        with open('coordinates.txt', 'w') as f:
            f.write(f'{x},{y}')
    if event == cv2.EVENT_LBUTTONDOWN:
        end_pos = (y, x)
        flag = True
    max_distance = math.sqrt((begin_pos[0] - end_pos[0]) ** 2 + (begin_pos[0] - end_pos[0]) ** 2)

def yolo2(frame):
    frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    frame = Image.fromarray(np.uint8(frame))
    frame = np.array(yolo.detect_image(frame))
    frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
    return frame

def unet2(frame):
    name_classes = ["background", "maze", "robot"]
    frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    frame = Image.fromarray(np.uint8(frame))
    frame = np.array(unet.detect_image(frame, count=True, name_classes=name_classes))
    frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
    return frame

def get_direction(begin_pos, next_node):
    x_diff = begin_pos[1] - next_node[1]
    y_diff = begin_pos[0] - next_node[0]
    angle = math.degrees(round(math.atan2(y_diff, x_diff), 2))
    return angle

def move(target_position, angle):
    r = 0.5
    target_position[0] += r * math.cos(-angle)
    target_position[1] += r * math.sin(-angle)
    return target_position

# ===========================
# Main Application Loop
# ===========================
choose = 0  
route = []
route_ = []
if __name__ == "__main__":
    cap = cv2.VideoCapture(0)  # Real-time input of ultrasound images from the data acquisition card
    ri = 1
    if not cap.isOpened():
        print("Failure to obtain ultrasound images")
    else:
        print("Successfully read the ultrasound image")

    size = (int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)), int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT)))
    size_change = (size[0] // ri, size[1] // ri)
    fourcc = cv2.VideoWriter_fourcc(*'XVID')
    output_video = cv2.VideoWriter('output_video2.avi', fourcc, 20.0, size_change)
    
    cv2.namedWindow("video", cv2.WINDOW_NORMAL)
    cv2.resizeWindow("video", size[0]//ri, size[1]//ri)
    cv2.namedWindow("video1", cv2.WINDOW_NORMAL)
    cv2.resizeWindow("video1", size[0]//ri, size[1]//ri)
    cv2.setMouseCallback("video1", event_handler)

    yolo = YOLO()
    unet = Unet()
    frame_count = 0

    while True:
        ret, frame = cap.read()
        print(frame.shape)
        frame_resized = cv2.resize(frame, size_change)
        frame_resized = unet2(frame_resized)

        with open('coordinates.txt', 'r') as f:
            line = f.readline()
            if line:
                center_x, center_y = map(int, line.split(','))
        begin_pos = (center_y, center_x)

        white_img = np.ones_like(frame_resized) * 255
        top_left_x = max(center_x - 85, 0)
        top_left_y = max(center_y - 85, 0)
        bottom_right_x = min(center_x + 85, frame_resized.shape[1])
        bottom_right_y = min(center_y + 85, frame_resized.shape[0])
        white_img[top_left_y:bottom_right_y, top_left_x:bottom_right_x] = frame_resized[top_left_y:bottom_right_y, top_left_x:bottom_right_x]

        frame1 = yolo2(white_img)

        pos.append(begin_pos)
        for row, (x, y) in enumerate(pos, start=1):
            ws1.cell(row=row, column=1, value=y)
            ws1.cell(row=row, column=2, value=x)
        wb1.save("pos.xlsx")
        print("Save table 1 successfully")

        gray_img = cv2.cvtColor(frame_resized, cv2.COLOR_RGB2GRAY)
        _, gray_img = cv2.threshold(gray_img, 50, 255, cv2.THRESH_BINARY_INV)
        kernel = np.ones((3, 3), np.uint8)
        gray_img = cv2.dilate(gray_img, kernel, iterations=10)

        gray_img[max(0, center_y-30):min(gray_img.shape[0], center_y+30), max(0, center_x-30):min(gray_img.shape[1], center_x+30)] = 0

        # ------------------- BFS Path Planning ----------------------
        if choose == 0:
            kernel_1 = np.ones((2, 2), np.uint8)
            gray_img_ = cv2.dilate(gray_img, kernel_1)
            Matrix = gray_img_
            route = []
            route_ = []
            length = Matrix.shape[0]
            width = Matrix.shape[1]
            Maxlength = Matrix.shape[0] * Matrix.shape[1]
            front = np.zeros((Maxlength, 1))
            visit = np.ones((Maxlength, 1))

            if frame_count % 1 == 0:
                if BFS(Matrix, begin_pos, end_pos, length, front, visit):
                    begin = end_pos[0] + end_pos[1] * length
                    while begin != 0:
                        route.append([begin % length, int(begin / length)])
                        route_.append(begin)
                        begin = int(front[begin][0])
                    route = [[index % length, index // length] for index in reversed(route_)]
                    # Fit a Bezier curve to smooth the path and fits the bladder wall
                    if len(route) >= 4:
                        route = bezier_curve(route, n_points=len(route) * 3)
                        next_node = route[10]
                    else:
                        next_node = route[-1]

                    for row, (x, y) in enumerate(pos_next, start=1):
                        ws2.cell(row=row, column=1, value=y)
                        ws2.cell(row=row, column=2, value=x)
                    wb2.save("pos_next.xlsx")
                    print("Save table 2 successfully")

                    direction = get_direction(begin_pos, next_node)
                    print(direction)

                    with open(fname, 'r+', encoding='utf-8') as f:
                        target_position = list(map(float, f.readline().strip().split(',')))

                    print(cps.moveL(target_position))
                    cps.waitMoveDone()
                    print(f"The robot moves to the coordinates: {target_position}")

                    target_position = move(target_position, direction)
                    with open(fname, 'w', encoding='utf-8') as f:
                        line = ','.join(map(str, target_position)) + '\n'
                        f.write(line)

                print(f"数组已保存到 {fname} 文件中")

        for i, j in enumerate(route[1::]):
            gray_img[route_[i] % length][int(route_[i] / length)] = 255

        render(route, frame_resized)
        kernel_1 = np.ones((2, 2), np.uint8)
        gray_img = cv2.dilate(gray_img, kernel_1)
        cv2.imshow("video", gray_img)

        frame_count += 1
        if cv2.waitKey(1) & 0xFF == 27:
            break

    cap.release()
    output_video.release()
    cv2.destroyAllWindows()
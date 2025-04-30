import cv2
import numpy as np
import tensorflow as tf
from PIL import Image
from yolo import YOLO, YOLO_ONNX
from collections import deque
from unet import Unet
import time
import math
import matplotlib.pyplot as plt
from collections import Counter
from mss import mss
from collections import deque
import openpyxl
import serial

gpus = tf.config.experimental.list_physical_devices(device_type='GPU')
for gpu in gpus:
    tf.config.experimental.set_memory_growth(gpu, True)


wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = 'pos1.xlsx'

wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'pos_next1.xlsx'


last_pos = (0, 0)
begin_pos = (0, 0)
end_pos = (0, 10)
pos=[]
pos_next=[]
flag = True
sct = mss()

monitor = sct.monitors[2]  

cv2.namedWindow('Screen Capture', cv2.WINDOW_NORMAL)

cv2.resizeWindow('Screen Capture', monitor["width"], monitor["height"])

def render(route, frame):
    for i in range(len(route) - 1):

        x1, y1 = route[i]
        x2, y2 = route[i + 1]
        

        cv2.line(frame, (y1, x1), (y2, x2), (0, 0, 255), 2)

    cv2.putText(frame, f"Mouse: {last_pos}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)
    
    cv2.imshow("video1", frame)
    
    cv2.waitKey(1)
def event_handler(event, x, y, flags, param):
    global last_pos
    global begin_pos
    global end_pos

    if event == cv2.EVENT_RBUTTONDOWN:
        center_x, center_y = x, y
        with open('coordinates.txt', 'w') as f:
           f.write(f'{center_x},{center_y}')
           
    if event == cv2.EVENT_LBUTTONDOWN:
        last_pos = (y, x)
        end_pos = last_pos
        

def BFS_choice(cur, Matrix, front, visit, q):
    """方向选择"""
    Maxlength = Matrix.shape[0] * Matrix.shape[1]
    length = Matrix.shape[0]
    width = Matrix.shape[1]
    directions = [(1, 0), (-1, 0), (0, 1), (0, -1), (1, 1), (-1, -1), (1, -1), (-1, 1)]  

    for dx, dy in directions:
        new_x = cur % length + dx
        new_y = cur // length + dy

        if 0 <= new_x < length and 0 <= new_y < width:
            new_index = new_x + new_y * length
            if visit[new_index] and not Matrix[new_x][new_y]:
                front[new_index] = cur
                visit[new_index] = 0
                q.append(new_index)


def BFS(Matrix, begin_pos, target, length, front, visit):
    target_flag = False
    target_index = target[0] + target[1] * Matrix.shape[0]
    begin_index = begin_pos[0] + begin_pos[1] * Matrix.shape[0]
    q = deque()
    q.append(begin_index)
    visit[begin_index] = 0
    while q:
        cur = q.popleft()
        if target_index == cur:
            target_flag = True
            break
        BFS_choice(cur, Matrix, front, visit, q)
    return target_flag


def yolo2(frame):
    frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    frame = Image.fromarray(np.uint8(frame))
    frame = np.array(yolo.detect_image(frame))
    frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
    return frame

def unet2(frame):
    name_classes    = ["background","maze","robot"]
    frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    frame = Image.fromarray(np.uint8(frame))
    frame = np.array(unet.detect_image(frame, count=True, name_classes=name_classes))
    frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
    return frame

def get_direction(begin_pos, next_node):
    x_diff =  (begin_pos[1] - next_node[1])
    y_diff =  begin_pos[0] - next_node[0]

    # angle = round(math.atan2(y_diff, x_diff), 2)
    angle = -(math.degrees(round(math.atan2(y_diff, x_diff), 2)))
    # angle=angle+180

    bata = 90.00
    frequency = 1
    with open('angel.txt', 'w') as f:
         f.write(f"{angle:.2f}") 
    ser = serial.Serial('COM4', 115200) 
    
    angle_bytes = str(angle).encode()
    bata_bytes = str(bata).encode()
    frequency_bytes = str(frequency).encode()
    data = b",".join([angle_bytes, bata_bytes, frequency_bytes])+b'\n'
    ser.write(data)
    
    return angle



if __name__ == "__main__":
    yolo = YOLO()
    unet = Unet()
    ri=1
    mi=4
    size=(monitor["width"], monitor["height"])
    size_change=(size[0]//ri, size[1]//ri)
    size_change1=(size[0]//mi, size[1]//mi)
    fourcc = cv2.VideoWriter_fourcc(*'XVID')
    output_video = cv2.VideoWriter('move.avi', fourcc, 5.0, size_change)
    output_video.set(cv2.CAP_PROP_BITRATE,20000)
    cv2.namedWindow("video", cv2.WINDOW_NORMAL)
    cv2.resizeWindow("video", size[0]//mi, size[1]//mi)
    cv2.namedWindow("video1", cv2.WINDOW_NORMAL)
    cv2.resizeWindow("video1",size[0]//mi, size[1]//mi)
    cv2.setMouseCallback("video1", event_handler)
    frame_count=0
    while True:
        t1 = time.time()
        screenshot = sct.grab(monitor)
        frame = np.array(screenshot)
        frame = cv2.cvtColor(frame, cv2.COLOR_BGRA2BGR)
        frame_resized = cv2.resize(frame, size_change1)
        frame_resized = unet2(frame_resized)
        frame_resized1 = cv2.resize(frame, size_change)
        output_video.write(frame_resized1)
        with open('coordinates.txt', 'r') as f:
            line = f.readline()
            if line:
                center_x, center_y = map(int, line.split(','))

        begin_pos = (center_y , center_x )
        if begin_pos:
                white_img = np.ones_like(frame_resized) * 255

                center_y, center_x = begin_pos
                width, height = 20, 20 

                top_left_x = max(center_x - width // 2, 0)
                top_left_y = max(center_y - height // 2, 0)
                bottom_right_x = min(center_x + width // 2, frame_resized.shape[1])
                bottom_right_y = min(center_y + height // 2, frame_resized.shape[0])

                white_img[top_left_y:bottom_right_y, top_left_x:bottom_right_x] = frame_resized[top_left_y:bottom_right_y, top_left_x:bottom_right_x]

                frame1 = yolo2(white_img)
        with open('coordinates.txt', 'r') as f:
            line = f.readline()
            if line:
                center_x, center_y = map(int, line.split(','))
        pos.append(begin_pos)
        for row, (x, y) in enumerate(pos, start=1):
            ws1.cell(row=row, column=1, value=y)
            ws1.cell(row=row, column=2, value=x)
        wb1.save("pos1.xlsx")
        print("保存表1成功")
        begin_pos = (center_y , center_x )
        
        gray_img = cv2.Canny(frame_resized, 100, 150)
        window_size=10
        min_x=max(0,center_x-window_size)
        max_x=min(gray_img.shape[1],center_x+window_size)
        min_y=max(0,center_y-window_size)
        max_y=min(gray_img.shape[0],center_y+window_size)
        gray_img[min_y:max_y,min_x:max_x]=0


        kernel_1 = np.ones((3, 3), np.uint8)
        gray_img_ = cv2.dilate(gray_img, kernel_1)
        
        Matrix = gray_img_
        route = []
        route_ = []
        length = Matrix.shape[0]
        width = Matrix.shape[1]
        Maxlength = Matrix.shape[0] * Matrix.shape[1]
        front = np.zeros((Maxlength,1))    
        visit = np.ones((Maxlength,1))     
        if frame_count % 1==0:
            if BFS(Matrix, begin_pos, end_pos, length, front, visit):
                begin = end_pos[0] + end_pos[1] * length
                while begin != 0:
                    route.append([begin % length, int(begin / length)])
                    route_.append(begin)
                    begin = int(front[begin][0])
                    route = [[index % length, index // length] for index in reversed(route_)] 
                # route.reverse()
                next_node=route[2]
                pos_next.append(next_node)
            # if len(pos) % 1 == 0:
                for row, (x, y) in enumerate(pos_next, start=1):
                    ws2.cell(row=row, column=1, value=y)
                    ws2.cell(row=row, column=2, value=x)

                # 保存 Excel 文件
                wb2.save("pos_next1.xlsx")
                print("保存表2成功")
                # 计算起点到下一个节点的运动方向
                direction = get_direction(begin_pos, next_node)
                print("起点到下一个节点要走的方向编码：", direction)
            else:
                print("失败") 
        
        for i, j in enumerate(route[1::]):
            gray_img[route_[i]%length][int(route_[i]/length)] = 255  
        
        render(route, frame_resized)
        # render1(route, frame_resized1)
        kernel_1 = np.ones((2, 2), np.uint8)
        gray_img = cv2.dilate(gray_img, kernel_1)
        # frame_1 = cv2.resize(gray_img , (640 , 480))
        cv2.imshow("video",(gray_img)) 
        frame_count+=1
        if cv2.waitKey(1) & 0xFF == 27:  # ESC 键的 ASCII 码为 27
           break

    # 释放窗口资源
    output_video.release()
    cv2.destroyAllWindows()
           